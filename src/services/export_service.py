import csv
import json
import os
import re
from html.parser import HTMLParser
from pathlib import Path

from docx import Document

from app.paths import json_for_text
from services.text_service import prepare_text_for_tei, tag_entities_tei


def export_txt(scan_files: list[dict], target_path: str) -> None:
    merged_content = []

    for pair in scan_files:
        txt_path = pair["txt"]
        if os.path.exists(txt_path):
            with open(txt_path, "r", encoding="utf-8") as handle:
                text_content = handle.read().strip()
                if text_content:
                    merged_content.append(text_content)

    with open(target_path, "w", encoding="utf-8") as handle:
        handle.write("\n\n".join(merged_content))


def _find_html_tables(text: str) -> list[tuple[str, str]]:
    return re.findall(r"<table\b([^>]*)>(.*?)</table>", text, flags=re.IGNORECASE | re.DOTALL)


def _find_html_table_elements(text: str) -> list[str]:
    return [
        match.group(0).strip()
        for match in re.finditer(r"<table\b[^>]*>.*?</table>", text, flags=re.IGNORECASE | re.DOTALL)
        if match.group(0).strip()
    ]


def _iter_html_table_matches(text: str):
    return re.finditer(r"<table\b[^>]*>.*?</table>", text, flags=re.IGNORECASE | re.DOTALL)


def _text_without_html(value: str) -> str:
    value = re.sub(r"<\s*br\s*/?\s*>", "\n", value, flags=re.IGNORECASE)
    value = re.sub(r"</\s*p\s*>", "\n\n", value, flags=re.IGNORECASE)
    value = re.sub(r"<[^>]+>", " ", value)
    return "\n".join(" ".join(line.split()) for line in value.splitlines()).strip()


def _caption_before_table(text: str, table_start: int) -> str:
    prefix = text[:table_start]
    previous_table_end = prefix.lower().rfind("</table>")
    if previous_table_end >= 0:
        prefix = prefix[previous_table_end + len("</table>"):]

    prefix = _text_without_html(prefix).strip()
    if not prefix:
        return ""

    matches = list(re.finditer(r"(?:^|\n)\s*\d+[A-Za-z]?\.\s+\S", prefix))
    if not matches:
        return ""

    caption = prefix[matches[-1].start():].strip()
    caption = re.sub(r"\s+", " ", caption).strip()
    if len(caption) < 4:
        return ""
    return caption


def _positive_int(value: str | None, default: int = 1) -> int:
    try:
        parsed = int(value or default)
    except (TypeError, ValueError):
        return default
    return max(parsed, 1)


class _HtmlTableParser(HTMLParser):
    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.rows: list[list[str]] = []
        self.merges: list[tuple[int, int, int, int]] = []
        self._row_index = -1
        self._col_index = 0
        self._occupied: set[tuple[int, int]] = set()
        self._cell_parts: list[str] | None = None
        self._cell_start: tuple[int, int, int, int] | None = None

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        tag = tag.lower()
        if tag == "tr":
            self._row_index += 1
            self._col_index = 0
            self._ensure_cell(self._row_index, 0)
            return

        if tag not in {"td", "th"} or self._row_index < 0:
            return

        attrs_dict = dict(attrs)
        while (self._row_index, self._col_index) in self._occupied:
            self._col_index += 1

        rowspan = _positive_int(attrs_dict.get("rowspan"))
        colspan = _positive_int(attrs_dict.get("colspan"))
        self._cell_start = (self._row_index, self._col_index, rowspan, colspan)
        self._cell_parts = []

    def handle_data(self, data: str) -> None:
        if self._cell_parts is not None:
            self._cell_parts.append(data)

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if tag not in {"td", "th"} or self._cell_parts is None or self._cell_start is None:
            return

        row, col, rowspan, colspan = self._cell_start
        value = " ".join("".join(self._cell_parts).split())
        self._ensure_cell(row + rowspan - 1, col + colspan - 1)
        self.rows[row][col] = value

        for row_offset in range(rowspan):
            for col_offset in range(colspan):
                target = (row + row_offset, col + col_offset)
                if target != (row, col):
                    self._occupied.add(target)

        if rowspan > 1 or colspan > 1:
            self.merges.append((row + 1, col + 1, row + rowspan, col + colspan))

        self._col_index = col + colspan
        self._cell_parts = None
        self._cell_start = None

    def _ensure_cell(self, row: int, col: int) -> None:
        while len(self.rows) <= row:
            self.rows.append([])
        for row_values in self.rows:
            while len(row_values) <= col:
                row_values.append("")


def _parse_html_table(table_html: str) -> tuple[list[list[str]], list[tuple[int, int, int, int]]]:
    parser = _HtmlTableParser()
    parser.feed(table_html)
    parser.close()
    rows = [row for row in parser.rows if any(cell.strip() for cell in row)]
    return rows, parser.merges


def _xlsx_sheet_title(base_title: str, used_titles: set[str]) -> str:
    title = re.sub(r"[:\\/?*\[\]]", "_", base_title).strip() or "table"
    title = title[:31]
    candidate = title
    suffix = 1
    while candidate in used_titles:
        suffix_text = f"_{suffix}"
        candidate = f"{title[:31 - len(suffix_text)]}{suffix_text}"
        suffix += 1
    used_titles.add(candidate)
    return candidate


def _html_document_for_table(table_html: str, title: str) -> str:
    return "\n".join(
        [
            "<!doctype html>",
            '<html lang="pl">',
            "<head>",
            '  <meta charset="utf-8">',
            f"  <title>{title}</title>",
            "</head>",
            "<body>",
            table_html,
            "</body>",
            "</html>",
            "",
        ]
    )


def export_html_tables_as_files(scan_files: list[dict]) -> int:
    written_count = 0

    for pair in scan_files:
        txt_path = Path(pair["txt"])
        if not txt_path.exists():
            continue

        text = txt_path.read_text(encoding="utf-8")
        tables = _find_html_table_elements(text)
        for index, table_html in enumerate(tables, start=1):
            target_path = txt_path.with_name(f"{txt_path.stem}-{index}.html")
            target_path.write_text(
                _html_document_for_table(table_html, target_path.stem),
                encoding="utf-8",
            )
            written_count += 1

    if not written_count:
        raise ValueError("Nie znaleziono tabel HTML w transkrypcjach.")

    return written_count


def export_html_tables_xlsx(scan_files: list[dict], target_path: str) -> int:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    workbook = Workbook()
    workbook.remove(workbook.active)
    used_titles = set()
    table_count = 0

    for pair in scan_files:
        txt_path = Path(pair["txt"])
        if not txt_path.exists():
            continue

        text = txt_path.read_text(encoding="utf-8")
        for index, match in enumerate(_iter_html_table_matches(text), start=1):
            table_html = match.group(0).strip()
            rows, merges = _parse_html_table(table_html)
            if not rows:
                continue

            sheet_title = _xlsx_sheet_title(f"{txt_path.stem}-{index}", used_titles)
            worksheet = workbook.create_sheet(sheet_title)
            for row_index, row in enumerate(rows, start=1):
                for col_index, value in enumerate(row, start=1):
                    worksheet.cell(row=row_index, column=col_index, value=value)
            for start_row, start_col, end_row, end_col in merges:
                worksheet.merge_cells(
                    start_row=start_row,
                    start_column=start_col,
                    end_row=end_row,
                    end_column=end_col,
                )
            caption = _caption_before_table(text, match.start())
            if caption:
                caption_row = len(rows) + 2
                caption_cell = worksheet.cell(row=caption_row, column=1, value=caption)
                caption_cell.font = Font(italic=True)
                caption_cell.alignment = Alignment(wrap_text=True, vertical="top")
                max_columns = max(len(row) for row in rows)
                if max_columns > 1:
                    worksheet.merge_cells(
                        start_row=caption_row,
                        start_column=1,
                        end_row=caption_row,
                        end_column=max_columns,
                    )
            table_count += 1

    if not table_count:
        raise ValueError("Nie znaleziono tabel HTML w transkrypcjach.")

    workbook.save(target_path)
    return table_count


def _find_table_header(table_body: str) -> str:
    match = re.search(r"<thead\b[^>]*>.*?</thead>", table_body, flags=re.IGNORECASE | re.DOTALL)
    return match.group(0).strip() if match else ""


def _find_table_rows(table_body: str) -> list[str]:
    body_without_headers = re.sub(r"<thead\b[^>]*>.*?</thead>", "", table_body, flags=re.IGNORECASE | re.DOTALL)
    tbody_matches = re.findall(r"<tbody\b[^>]*>(.*?)</tbody>", body_without_headers, flags=re.IGNORECASE | re.DOTALL)
    row_sources = tbody_matches if tbody_matches else [body_without_headers]
    rows = []
    for source in row_sources:
        rows.extend(
            row.strip()
            for row in re.findall(r"<tr\b[^>]*>.*?</tr>", source, flags=re.IGNORECASE | re.DOTALL)
            if row.strip()
        )
    return rows


def export_merged_html_table(scan_files: list[dict], target_path: str) -> int:
    table_attrs = ""
    table_header = ""
    table_rows = []

    for pair in scan_files:
        txt_path = pair["txt"]
        if not os.path.exists(txt_path):
            continue

        with open(txt_path, "r", encoding="utf-8") as handle:
            tables = _find_html_tables(handle.read())

        for attrs, table_body in tables:
            if not table_attrs:
                table_attrs = attrs.strip()
            if not table_header:
                table_header = _find_table_header(table_body)
            table_rows.extend(_find_table_rows(table_body))

    if not table_rows:
        raise ValueError("Nie znaleziono wierszy tabel HTML w transkrypcjach.")

    table_open = f"<table {table_attrs}>".replace("  ", " ").strip() if table_attrs else "<table>"
    lines = [
        "<!doctype html>",
        '<html lang="pl">',
        "<head>",
        '  <meta charset="utf-8">',
        "</head>",
        "<body>",
        table_open,
    ]
    if table_header:
        lines.append(table_header)
    lines.append("<tbody>")
    lines.extend(table_rows)
    lines.append("</tbody>")
    lines.append("</table>")
    lines.append("</body>")
    lines.append("</html>")

    with open(target_path, "w", encoding="utf-8") as handle:
        handle.write("\n".join(lines))

    return len(table_rows)


def export_docx(scan_files: list[dict], target_path: str) -> None:
    doc = Document()

    for pair in scan_files:
        if os.path.exists(pair["txt"]):
            all_lines = []
            with open(pair["txt"], "r", encoding="utf-8") as handle:
                lines = handle.readlines()
                all_lines.extend([line.strip() for line in lines])

            full_text = ""
            for line in all_lines:
                if not line:
                    full_text += "\n\n"

                if full_text == "":
                    full_text = line
                else:
                    if full_text.endswith("-"):
                        full_text = full_text[:-1] + line
                    else:
                        full_text += " " + line

            doc.add_paragraph(full_text)

    doc.save(target_path)


def export_tei(scan_files: list[dict], target_path: str) -> None:
    tei_content = []
    tei_content.append('<?xml version="1.0" encoding="UTF-8"?>')
    tei_content.append('<TEI xmlns="http://www.tei-c.org/ns/1.0">')
    tei_content.append("  <teiHeader>")
    tei_content.append("    <fileDesc>")
    tei_content.append("      <titleStmt><title>Eksport z ScansAndTranscriptions</title></titleStmt>")
    tei_content.append("      <publicationStmt><p>Wygenerowano automatycznie</p></publicationStmt>")
    tei_content.append("      <sourceDesc><p>Transkrypcje skanów</p></sourceDesc>")
    tei_content.append("    </fileDesc>")
    tei_content.append("  </teiHeader>")
    tei_content.append("  <text>")
    tei_content.append("    <body>")

    for pair in scan_files:
        if not os.path.exists(pair["txt"]):
            continue

        with open(pair["txt"], "r", encoding="utf-8") as handle:
            raw_text = handle.read()

        entities = {}
        json_path = json_for_text(pair["txt"])
        if json_path.exists():
            with json_path.open("r", encoding="utf-8") as handle:
                entities = json.load(handle).get("entities", {})

        processed_text = prepare_text_for_tei(raw_text)
        tagged_text = tag_entities_tei(processed_text, entities)

        tei_content.append(f'      <div type="page" n="{pair["name"]}">')
        for paragraph in tagged_text.split("\n\n"):
            if paragraph.strip():
                tei_content.append(f"        <p>{paragraph.strip()}</p>")
        tei_content.append("      </div>")

    tei_content.append("    </body>")
    tei_content.append("  </text>")
    tei_content.append("</TEI>")

    with open(target_path, "w", encoding="utf-8") as handle:
        handle.write("\n".join(tei_content))


def collect_ner_rows(scan_files: list[dict]) -> list[dict]:
    rows = []

    for pair in scan_files:
        json_path = json_for_text(pair["txt"])
        txt_path = Path(pair["txt"])

        if json_path.exists() and txt_path.exists():
            with json_path.open("r", encoding="utf-8") as handle:
                entities = json.load(handle).get("entities", {})

            for category, names in entities.items():
                for name in names:
                    rows.append(
                        {
                            "orig": name,
                            "cat": category,
                            "file": os.path.basename(pair["img"]),
                        }
                    )

    return rows


def unique_ner_names(rows: list[dict]) -> list[str]:
    return sorted({row["orig"] for row in rows})


def write_ner_csv(
    target_path: str,
    rows: list[dict],
    nominative_map: dict,
    headers: list[str],
) -> None:
    with open(target_path, "w", encoding="utf-8", newline="") as csvfile:
        writer = csv.writer(csvfile, delimiter=";")
        writer.writerow(headers)

        for row in rows:
            base_name = nominative_map.get(row["orig"], row["orig"])
            writer.writerow([row["orig"], base_name, row["cat"], row["file"]])
